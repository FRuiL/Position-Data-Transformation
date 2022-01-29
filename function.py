import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font
import os




def data_collec():

    #path:输入数据文件路径(可以将文件直接拖入终端)
    path = input('将原始文件拖至此处:').strip()
    year = input('请输入报告期:')
    prefer = input('\n请选择仓位查看方法:\n1.持股总市值\n2.持股数量\n请输入序号:')
    inv_mkt = input('\n请选择市场:\n1.上海 2.深圳 3.香港联交所\n请输入序号(用空格分隔):').split()
    inv_ty = input(
                '\n请选择投资类型:\n1.REITs\n2.被动指数型基金\n3.股票多空\
                \n4.混合债券型二级基金\n5.混合债券型一级基金\n6.灵活配置型基金\
                \n7.偏股混合型基金\n8.偏债混合型基金\n9.平衡混合型基金\
                \n10.普通股票型基金\n11.增强指数型基金\n12.增强指数型债券基金\
                \n13.中长期纯债型基金\n请输入序号(用空格分隔):'
                ).split()

    #创建字典并收集input数据
    markets = {'1':'上海', '2':'深圳', '3':'香港联交所'}
    invest_types = {
                    '1':'REITs', '2':'被动指数型基金', '3':'股票多空',\
                    '4':'混合债券型二级基金','5':'混合债券型一级基金',\
                    '6':'灵活配置型基金','7':'偏股混合型基金','8':'偏债混合型基金',\
                    '9':'平衡混合型基金','10':'普通股票型基金','11':'增强指数型基金',\
                    '12':'增强指数型债券基金','13':'中长期纯债型基金'
                    }

    m = []
    for mkt in inv_mkt:
        if mkt in markets.keys():
            m.append(markets[mkt])

    i = []
    for ty in inv_ty:
        if ty in invest_types.keys():
            i.append(invest_types[ty])

    #创建字符串便于生成文件名
    str_m = "+".join(m)
    str_i = "+".join(i)
    str = str_m + '+' + str_i


    #按投资类型条件判断
    if prefer == '1':
        companies = pd.read_excel(path, header=0)

        #获取新excel路径
        (root, filename) = os.path.split(path)

        #获得新文件路径
        str1 = str + '+' + '持股总市值.xlsx'
        path1 = os.path.join(root,str1)

        #筛选数据：
        #市场：上海+深圳；
        #投资类型：灵活配置型基金+偏股混合型基金+平衡混合型基金+普通股票型基金
        companies = companies.loc[companies['市场'].isin(m)]
        companies = companies.loc[companies['投资类型'].isin(i)]

        #重仓股加总
        shares_num = companies['持股总市值(万元)'].sum()

        #银行板块加总
        companies = companies.loc[companies['行业'].isin(['银行'])]
        banks_num = companies['持股总市值(万元)'].sum()

        #整理不同基金所持同一银行数据
        companies_merge = companies.groupby(companies['股票简称'],\
        as_index=False).agg({'持股总市值(万元)':'sum'})

        #转置写入新文件
        a = pd.pivot_table(companies_merge,columns=[u'股票简称'],values=[u'持股总市值(万元)'])
        wb = openpyxl.load_workbook(path)
        a.to_excel(path1)


        #在新建文件中写入数据
        wb = openpyxl.load_workbook(path1)
        ws = wb.active
        ws.delete_cols(1)
        ws.insert_cols(1,3)

        ws.cell(1,2,'银行股市值(万元)')
        ws.cell(1,3,'重仓股总市值(万元)')
        ws.cell(2,1, year)
        ws.cell(2,2,banks_num)
        ws.cell(2,3,shares_num)
        ws.column_dimensions['B'].width=17.5
        ws.column_dimensions['C'].width=17.5
        #设置字体格式
        #font = Font(bold=True)
        #cell2 = ws["B1"]
        #cell2.font = font
        #cell3 = ws["C1"]
        #cell3.font = font
        wb.save(path1)

        print ('文件已生成。')


    elif prefer == '2':
        companies = pd.read_excel(path, header=0)

        #获取新excel路径
        (root, filename) = os.path.split(path)

        str2 = str + '+' + '持股数量.xlsx'
        path1 = os.path.join(root,str2)

        companies = companies.loc[companies['市场'].isin(m)]
        companies = companies.loc[companies['投资类型'].isin(i)]

        shares_num = companies['持股数量(万股)'].sum()

        companies = companies.loc[companies['行业'].isin(['银行'])]
        banks_num = companies['持股数量(万股)'].sum()

        companies_merge = companies.groupby(companies['股票简称'],\
        as_index=False).agg({'持股数量(万股)':'sum'})

        a = pd.pivot_table(companies_merge,columns=[u'股票简称'],values=[u'持股数量(万股)'])
        a.to_excel(path1)

        wb = openpyxl.load_workbook(path1)
        ws = wb.active
        ws.delete_cols(1)
        ws.insert_cols(1,3)

        ws.cell(1,2,'银行股股数(万股)')
        ws.cell(1,3,'重仓股总股数(万股)')
        ws.cell(2,1, year)
        ws.cell(2,2,banks_num)
        ws.cell(2,3,shares_num)
        ws.column_dimensions['B'].width=17.5
        ws.column_dimensions['C'].width=17.5
        wb.save(path1)

        print ('文件已生成。')

    else:
        print ('\n请输入有效数据.')
